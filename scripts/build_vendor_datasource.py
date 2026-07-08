#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook, load_workbook


PDF_PATH = Path("/Users/long/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/m5637379_954d/temp/drag/白名单商家信息表-上海汇总(1).pdf")
XLSX_PATH = Path("/Users/long/Downloads/白名单商家信息表-杭州汇总(2).xlsx")
OUT_DIR = Path("/Users/long/Documents/Codex/2026-07-08/yo/outputs")

FIELDS = [
    "record_id",
    "record_type",
    "city",
    "official_website",
    "official_phone",
    "official_email",
    "official_wechat",
    "showroom_locations",
    "service_area",
    "scope",
    "budget_positioning",
    "design_style",
    "materials",
    "environmental_standard",
    "typical_timeline",
    "verification_date",
    "verified_by",
    "vendor_label",
    "vendor_name",
    "contact_person",
    "address",
    "district",
    "store_type",
    "authorization",
    "customer_segments",
    "board_materials",
    "recommended_boards",
    "hardware",
    "starting_price_projection",
    "edge_band_material",
    "edge_banding_process",
    "edge_banding_equipment",
    "cabinet_structure",
    "cabinet_connectors_feet",
    "measurement_design_fee",
    "deposit_refund_policy",
    "designer_team",
    "designer_income",
    "rendering_cycle",
    "outsourced_design",
    "installation_cost_per_projection",
    "installation_team",
    "damage_handling",
    "floor_protection",
    "aftersales_staff",
    "warranty_policy",
    "customization_cycle",
    "retain_final_payment",
    "payment_terms",
    "creator_comment",
    "strengths",
    "caveats",
    "source_name",
    "source_type",
    "source_path_or_url",
    "source_author",
    "source_date",
    "source_page_or_rows",
    "source_confidence",
    "raw_text",
]

FIELD_DICTIONARY = {
    "record_id": "稳定记录 ID，后续引用和去重用。",
    "record_type": "vendor 表示商家记录；external_note 表示外部资料/笔记。",
    "city": "主要服务城市。",
    "official_website": "商家官网，来自公开来源。",
    "official_phone": "商家公开电话，来自官网等公开来源。",
    "official_email": "商家公开邮箱，来自官网等公开来源。",
    "official_wechat": "商家公开微信号，来自官网等公开来源。",
    "showroom_locations": "公开展示空间、工作室或展厅地址。",
    "service_area": "推荐 skill 使用的服务范围字段，包含城市、区、地址。",
    "scope": "适合业务范围，如全屋定制、柜体定制、整案设计、硬装、软装。",
    "budget_positioning": "按投影起价和资料描述粗略归纳的预算档位。",
    "design_style": "适合的风格标签，暂留给后续案例/小红书资料补充。",
    "materials": "推荐 skill 使用的材料字段，通常等于板材品类。",
    "environmental_standard": "从材料/授权中提取的环保或品牌关键词。",
    "typical_timeline": "推荐 skill 使用的交付周期字段。",
    "verification_date": "资料核验日期，原始文件未提供时留空。",
    "verified_by": "资料来源或核验人。",
    "vendor_label": "原始资料中的门店编号/标签。",
    "vendor_name": "正式商家名，当前资料多未提供，建议后续人工补充。",
    "contact_person": "资料中的联系人。",
    "address": "门店或工作室地址。",
    "district": "从地址推断的区县。",
    "creator_comment": "大 V 点评、建议和资料中的判断性备注。",
    "strengths": "根据资料自动推断的优势标签，需人工复核。",
    "caveats": "根据资料自动推断的提醒项，需人工复核。",
    "source_confidence": "结构化抽取置信度；PDF 通常 medium，XLSX 通常 high。",
    "raw_text": "原始抽取文本，便于回溯和人工校对。",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip()


def derive_district(address: str) -> str:
    match = re.search(r"上海市?([^市区县]{1,8}[区县])", address)
    if match:
        return match.group(1)
    match = re.search(r"杭州市?([^市区县]{1,8}[区县])", address)
    if match:
        return match.group(1)
    if "滨江区" in address:
        return "滨江区"
    return ""


def base_record(**kwargs: str) -> dict[str, str]:
    record = {field: "" for field in FIELDS}
    record.update({key: clean(value) for key, value in kwargs.items()})
    return record


def price_position(value: str) -> str:
    numbers = [int(item) for item in re.findall(r"\d+", value)]
    if not numbers:
        return ""
    price = max(numbers)
    if price >= 1500:
        return "中高预算/高配置"
    if price >= 1200:
        return "中高预算"
    if price >= 1000:
        return "中预算/性价比"
    return "性价比/预算友好"


def environmental_terms(value: str) -> str:
    hits = []
    for term in ["ENF", "HENF", "F4星", "爱格", "可丽芙", "菲德莱", "靓时", "水性科天", "福人", "兔宝宝"]:
        if term.lower() in value.lower():
            hits.append(term)
    return "；".join(dict.fromkeys(hits))


def enrich_record(record: dict[str, str]) -> dict[str, str]:
    if record["record_type"] != "vendor":
        return record
    area_parts = [record["city"], record["district"], record["address"]]
    record["service_area"] = " ".join(part for part in area_parts if part)
    scope_parts = ["全屋定制", "柜体定制"]
    if "整案" in record["store_type"] or "设计" in record["store_type"] or "全案" in record["creator_comment"]:
        scope_parts.append("整案/设计服务")
    if "硬装" in record["store_type"]:
        scope_parts.append("硬装")
    if "软装" in record["store_type"]:
        scope_parts.append("软装")
    record["scope"] = "；".join(dict.fromkeys(scope_parts))
    record["budget_positioning"] = price_position(record["starting_price_projection"])
    record["materials"] = record["board_materials"]
    record["environmental_standard"] = environmental_terms(record["board_materials"] + " " + record["authorization"])
    record["typical_timeline"] = record["customization_cycle"]
    record["verified_by"] = "猴哥/大V白名单资料"
    if record["record_id"] == "HZ-004":
        official_note = (
            "官网补充：澎升设计 Pengsheng Design，定位为杭州本土化设计公司，强调私宅设计、产品设计、"
            "软装与全屋定制服务；官网列出木作/全屋定制、拱墅工作室和临平全屋定制木作展厅。"
        )
        record["vendor_name"] = "澎升设计 Pengsheng Design"
        record["official_website"] = "https://pengshengdesign.com"
        record["official_phone"] = "+86 13735800742"
        record["official_email"] = "m5637379@gmail.com"
        record["official_wechat"] = "PengshengDesign"
        record["address"] = "杭州市拱墅区影天像素园澎升设计"
        record["district"] = "拱墅区"
        record["showroom_locations"] = (
            "拱墅工作室：杭州市拱墅区影天像素园澎升设计；"
            "临平全屋定制木作展厅：杭州市临平区元星大厦澎升设计"
        )
        record["service_area"] = (
            "杭州 拱墅区 临平区；拱墅工作室；临平全屋定制木作展厅；"
            "杭州市拱墅区影天像素园澎升设计；杭州市临平区元星大厦澎升设计"
        )
        record["scope"] = "全屋定制；柜体定制；整案/设计服务；私宅设计；软装；全屋定制木作"
        record["store_type"] = "设计公司/工作室/木作展厅"
        record["verification_date"] = "2026-07-08"
        record["creator_comment"] = clean(record["creator_comment"] + "；" + official_note)
        record["source_name"] = clean(record["source_name"] + "；pengshengdesign.com 官网")
        record["source_path_or_url"] = clean(record["source_path_or_url"] + "；https://pengshengdesign.com；https://pengshengdesign.com/contact；https://pengshengdesign.com/millwork")
    return record


def first_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.S)
        if match:
            return clean(match.group(1))
    return ""


def split_contact_and_comment(line: str) -> tuple[str, str]:
    match = re.match(r"^(.{1,8}?(?:老师|老板|总|姐|哥|经理|先生|女士))\s*(.*)$", line)
    if match:
        return clean(match.group(1)), clean(match.group(2))
    parts = line.split(" ", 1)
    if len(parts) == 2:
        return clean(parts[0]), clean(parts[1])
    return clean(line), ""


def pdf_authorization(text: str) -> str:
    value = first_match(text, [r"授权体系\s+(.+?)\s*核心客群"])
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        if clean(line) == "授权体系":
            before = clean(lines[idx - 1]) if idx > 0 else ""
            after = clean(lines[idx + 1]) if idx + 1 < len(lines) else ""
            if before and after and not before.startswith(("门店", "一、", "二、", "三、", "四、")):
                return clean(before + after)
    return value


def pdf_blocks() -> list[tuple[int, str, str]]:
    chunks: list[tuple[int, str, str]] = []
    with pdfplumber.open(PDF_PATH) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            match = re.search(r"(门店[一二三四五六七八九十]+（[^）]+）)", text)
            label = match.group(1) if match else f"门店{page_no}"
            chunks.append((page_no, label, text))
    return chunks


def extract_pdf_record(page_no: int, label: str, text: str) -> dict[str, str]:
    comment = first_match(text, [r"猴哥点评/建议\s*(.*?)\s*门店联系人"])
    contact_line = first_match(text, [r"门店联系人\s+([^\n]+)"])
    contact_person, contact_comment = split_contact_and_comment(contact_line)
    if contact_comment and contact_comment not in comment:
        comment = clean(" ".join([comment, contact_comment]))
    record = base_record(
        record_type="vendor",
        city="上海",
        vendor_label=label,
        contact_person=contact_person,
        address=first_match(text, [r"门店地址\s+([^\n]+)"]),
        store_type=first_match(text, [r"门店类型\s+([^\n]+)"]),
        authorization=pdf_authorization(text),
        customer_segments=first_match(text, [r"核心客群\s+(.+?)\s+服务能力"]),
        board_materials=first_match(text, [r"板材品类\s+(.+?)\s+五金品类"]),
        recommended_boards=first_match(text, [r"猴哥推荐搭配（基于板材物理性能和饰面美观度考虑）\s*(.+?)\s*五金品类"]),
        hardware=first_match(text, [r"五金品类\s+(.+?)\s+投影面积起价"]),
        starting_price_projection=first_match(text, [r"投影面积起价\s+([^\n]+)"]),
        edge_band_material=first_match(text, [r"（PVC/ABS/PP）\s+([^\n]+)"]),
        edge_banding_process=first_match(text, [r"（激光 / PUR / EVA）\s+([^\n]+)"]),
        edge_banding_equipment=first_match(text, [r"封边设备\s+(.+?)\s+柜体工艺"]),
        cabinet_structure=first_match(text, [r"柜体工艺\s+(.+?)\s+柜体\s*（单元柜/非单元柜）", r"（单元柜/非单元柜）\s+(.+?)\s+柜体连接件"]),
        cabinet_connectors_feet=first_match(text, [r"柜体连接件、调整脚\s+(.+?)\s+四、服务信息", r"柜体连接件、调整脚\s+([^\n]+)"]),
        measurement_design_fee=first_match(text, [r"测量设计免费 or 收费\s+(.+?)\s+定金是否可退"]),
        deposit_refund_policy=first_match(text, [r"定金是否可退\s+([^\n]+)"]),
        designer_team=first_match(text, [r"设计人员配置\s+([^\n]+)"]),
        designer_income=first_match(text, [r"设计师人均月收入\s+([^\n]+)"]),
        rendering_cycle=first_match(text, [r"效果图出图周期\s+(.+?)\s+是否使用外包设计师"]),
        outsourced_design=first_match(text, [r"是否使用外包设计师\s+([^\n]+)"]),
        installation_cost_per_projection=first_match(text, [r"安装工人工资水平（每投影安装成本）\s+([^\n]+)"]),
        installation_team=first_match(text, [r"安装团队配置\s+([^\n]+)"]),
        damage_handling=first_match(text, [r"（受损板材重新生产/补漆）\s+(.+?)\s+安装前是否进行地面保护"]),
        floor_protection=first_match(text, [r"安装前是否进行地面保护\s+(.+?)\s+是否有专职售后人员"]),
        aftersales_staff=first_match(text, [r"是否有专职售后人员\s+([^\n]+)"]),
        warranty_policy=first_match(text, [r"质保维修政策\s+([^\n]+)"]),
        customization_cycle=first_match(text, [r"定制周期需要多久？逾期赔付标准？\s+(.+?)\s+合同 是否留尾款", r"定制周期需要多久？逾期赔付标准？\s+(.+?)\s+是否留尾款"]),
        retain_final_payment=first_match(text, [r"是否留尾款？\s+([^\n]+)"]),
        payment_terms=first_match(text, [r"付款方式/付款节点\s+([^\n]+)"]),
        creator_comment=comment,
        source_name="白名单商家信息表-上海汇总(1).pdf",
        source_type="pdf",
        source_path_or_url=str(PDF_PATH),
        source_page_or_rows=f"page {page_no}",
        source_confidence="medium",
        raw_text=text,
    )
    record["district"] = derive_district(record["address"])
    record["strengths"] = infer_strengths(record)
    record["caveats"] = infer_caveats(record)
    return record


def infer_strengths(record: dict[str, str]) -> str:
    parts: list[str] = []
    text = " ".join(record.values())
    if re.search(r"设计能力强|整案|全案|设计师", text):
        parts.append("设计/整案能力较强")
    if re.search(r"交付快|30天|15到25天|25天", text):
        parts.append("交付速度有优势")
    if re.search(r"性价比|1080|980|1199", text):
        parts.append("性价比或入门价格有优势")
    if re.search(r"专职售后|老板亲自|售后人员", text):
        parts.append("售后响应配置较明确")
    if re.search(r"单元柜|IF|豪迈|激光|PUR", text):
        parts.append("柜体/封边工艺配置较完整")
    return "；".join(dict.fromkeys(parts))


def infer_caveats(record: dict[str, str]) -> str:
    parts: list[str] = []
    text = " ".join(record.values())
    if re.search(r"不退|不可退", record.get("deposit_refund_policy", "")):
        parts.append("定金退款规则需提前确认")
    if re.search(r"60|75|45个工作日", record.get("customization_cycle", "")):
        parts.append("工期不算短，需确认逾期责任")
    if re.search(r"不留尾款|付清|全款", record.get("payment_terms", "") + record.get("retain_final_payment", "")):
        parts.append("付款节点偏前，合同边界要写清")
    if re.search(r"1500|高端|豪宅", text):
        parts.append("预算门槛可能偏高")
    return "；".join(dict.fromkeys(parts))


def parse_xlsx_records() -> list[dict[str, str]]:
    ws = load_workbook(XLSX_PATH, data_only=True).active
    rows = [[clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    starts = [i for i, row in enumerate(rows) if row and re.match(r"门店[一二三四五六七八九十]+", row[0])]
    starts.append(len(rows))
    records: list[dict[str, str]] = []
    for idx in range(len(starts) - 1):
        start, end = starts[idx], starts[idx + 1]
        block = rows[start:end]
        label = block[0][0]
        data: dict[str, str] = {}
        comments: list[str] = []
        raw_lines: list[str] = []
        for absolute_row, row in enumerate(block, start + 1):
            raw_lines.append(f"{absolute_row}: " + " | ".join(row))
            key = row[1] or row[0]
            value = row[2] if len(row) > 2 else ""
            comment = row[3] if len(row) > 3 else ""
            if value and key and not re.match(r"[一二三四]、", key):
                data[key] = value
            if comment and "猴哥点评/建议" not in comment:
                comments.append(comment)
        record = base_record(
            record_type="vendor",
            city="杭州",
            vendor_label=label,
            contact_person=data.get("门店联系人", ""),
            address=data.get("门店地址", ""),
            store_type=data.get("门店类型", ""),
            authorization=data.get("授权体系", ""),
            customer_segments=data.get("核心客群", ""),
            board_materials=data.get("板材品类", ""),
            recommended_boards=next((c for c in comments if "柜体板建议" in c), ""),
            hardware=data.get("五金品类", ""),
            starting_price_projection=data.get("投影面积起价", ""),
            edge_band_material=data.get("封边带材质 （PVC/ABS/PP）", ""),
            edge_banding_process=data.get("封边工艺 （激光 / PUR / EVA）", ""),
            edge_banding_equipment=data.get("封边设备", ""),
            cabinet_structure=data.get("柜体工艺 （单元柜/非单元柜）", ""),
            cabinet_connectors_feet=data.get("柜体连接件、 调整脚", ""),
            measurement_design_fee=data.get("测量设计免费 or 收费", ""),
            deposit_refund_policy=data.get("定金是否可退", ""),
            designer_team=data.get("设计人员配置", ""),
            designer_income=data.get("设计师人均月收入", ""),
            rendering_cycle=data.get("效果图出图周期", ""),
            outsourced_design=data.get("是否使用外包设计师", ""),
            installation_cost_per_projection=data.get("安装工人工资水平（每投影安装成本）", ""),
            installation_team=data.get("安装团队配置", ""),
            damage_handling=data.get("安装、 运输过程中的磕碰 （受损板材重新生产/补漆）", ""),
            floor_protection=data.get("安装前是否进行地面保护", ""),
            aftersales_staff=data.get("是否有专职售后人员", ""),
            warranty_policy=data.get("质保维修政策", ""),
            customization_cycle=data.get("定制周期需要多久？逾期赔付标准？", ""),
            retain_final_payment=data.get("是否留尾款？", ""),
            payment_terms=data.get("付款方式/付款节点", ""),
            creator_comment="；".join(dict.fromkeys(comments)),
            source_name="白名单商家信息表-杭州汇总(2).xlsx",
            source_type="xlsx",
            source_path_or_url=str(XLSX_PATH),
            source_page_or_rows=f"rows {start + 1}-{end}",
            source_confidence="high",
            raw_text="\n".join(raw_lines),
        )
        record["district"] = derive_district(record["address"])
        record["strengths"] = infer_strengths(record)
        record["caveats"] = infer_caveats(record)
        records.append(record)
    return records


def xhs_record() -> dict[str, str]:
    text = (
        "后续来了，增加两家重量级全屋定制探店记录。作者提到新增 DCEO 和马库住建工厂。"
        "评分权重：口碑5%、展厅5%、设计25%、价格/性价比25%、服务20%、工艺20%。"
        "案例条件：二手房，投影面积40~42㎡，中古风；柜体考虑国产或者进口二线品牌，柜门使用爱格。"
        "原文提示资料为个人主观感受，只可大致参考，不作为决策依据。"
    )
    return base_record(
        record_id="XHS-20240628-001",
        record_type="external_note",
        city="杭州",
        vendor_label="小红书探店笔记",
        vendor_name="DCEO；马库住建工厂",
        creator_comment=text,
        strengths="提供评分维度和权重；可作为用户自评/推荐排序参考；补充两个杭州候选商家线索",
        caveats="网页正文未包含图片表格明细；需补截图或原始表格后才能精确入库",
        source_name="后续来了，增加两家重量级全屋定制探店记录",
        source_type="xiaohongshu_url",
        source_path_or_url="https://www.xiaohongshu.com/explore/667e224b000000001c0270f9?xsec_token=ABfKEsYxm5nNBo9Zi4-Uzk-UQ8iq-H3rKuFSyPbukjQjM=&xsec_source=pc_user",
        source_author="呼伦贝尔的风",
        source_date="2024-06-28",
        source_confidence="medium",
        raw_text=text,
    )


def xhs_listed_vendor_records() -> list[dict[str, str]]:
    source_url = "https://www.xiaohongshu.com/explore/667e224b000000001c0270f9?xsec_token=ABfKEsYxm5nNBo9Zi4-Uzk-UQ8iq-H3rKuFSyPbukjQjM=&xsec_source=pc_user"
    shared_note = (
        "来自小红书笔记《后续来了，增加两家重量级全屋定制探店记录》。"
        "正文提到新增 DCEO 和马库住建工厂；图片/表格明细未完整结构化抽取。"
        "原文评分权重：口碑5%、展厅5%、设计25%、价格/性价比25%、服务20%、工艺20%。"
        "案例条件：二手房，投影面积40~42㎡，中古风；柜体考虑国产或者进口二线品牌，柜门使用爱格。"
    )
    rows = []
    for idx, name in enumerate(["DCEO"], 1):
        rows.append(
            base_record(
                record_id=f"XHS-20240628-V{idx:02d}",
                record_type="vendor",
                city="杭州",
                vendor_label="小红书博主列出商家",
                vendor_name=name,
                service_area="杭州",
                scope="全屋定制；柜体定制",
                creator_comment=shared_note,
                strengths="博主探店笔记明确列出；可作为杭州候选商家线索",
                caveats="缺少结构化价格、地址、材料、工艺、售后等字段；需补具体截图/原表或二次核验后再强推荐",
                source_name="后续来了，增加两家重量级全屋定制探店记录",
                source_type="xiaohongshu_url",
                source_path_or_url=source_url,
                source_author="呼伦贝尔的风",
                source_date="2024-06-28",
                source_confidence="low",
                raw_text=shared_note,
            )
        )
    return rows


def xhs_search_record() -> dict[str, str]:
    url = "https://www.xiaohongshu.com/search_result?keyword=%25E6%259D%25AD%25E5%25B7%259E%25E5%2585%25A8%25E5%25B1%258B%25E5%25AE%259A%25E5%2588%25B6%25E6%258E%2592%25E5%2590%258D&source=web_user_page"
    text = (
        "小红书搜索入口，关键词为“杭州全屋定制排名”。网页可见标题和筛选入口，但未暴露具体笔记卡片、作者、日期或正文。"
        "该来源适合作为后续检索入口，不适合作为单独商家推荐证据。后续可从该搜索页挑选具体笔记 URL、截图或原始表格后再拆分入库。"
    )
    return base_record(
        record_id="XHS-SEARCH-20260708-001",
        record_type="external_note",
        city="杭州",
        vendor_label="小红书搜索入口",
        vendor_name="",
        creator_comment=text,
        strengths="可持续发现杭州全屋定制排名/避坑/探店相关笔记；适合后续人工筛选具体来源",
        caveats="搜索页没有稳定披露具体笔记内容；不能直接用于推荐结论；需二次打开具体笔记或保存截图",
        source_name="杭州全屋定制排名 - 小红书搜索",
        source_type="xiaohongshu_search_url",
        source_path_or_url=url,
        source_date="2026-07-08",
        source_confidence="low",
        raw_text=text,
    )


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    records = [extract_pdf_record(*item) for item in pdf_blocks()]
    records.extend(parse_xlsx_records())
    records.extend(xhs_listed_vendor_records())
    records.append(xhs_record())
    records.append(xhs_search_record())
    for i, record in enumerate(records, 1):
        if not record["record_id"]:
            city_code = "SH" if record["city"] == "上海" else "HZ" if record["city"] == "杭州" else "SRC"
            record["record_id"] = f"{city_code}-{i:03d}"
    records = [enrich_record(item) for item in records]
    csv_path = OUT_DIR / "whole-home-customization-vendors.csv"
    json_path = OUT_DIR / "whole-home-customization-vendors.json"
    xlsx_path = OUT_DIR / "whole-home-customization-vendors.xlsx"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(records)
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "vendors"
    ws.append(FIELDS)
    for record in records:
        ws.append([record.get(field, "") for field in FIELDS])
    dict_ws = wb.create_sheet("field_dictionary")
    dict_ws.append(["field", "meaning"])
    for field, meaning in FIELD_DICTIONARY.items():
        dict_ws.append([field, meaning])
    wb.save(xlsx_path)
    summary = {
        "total_records": len(records),
        "vendor_records": sum(1 for item in records if item["record_type"] == "vendor"),
        "external_note_records": sum(1 for item in records if item["record_type"] == "external_note"),
        "cities": sorted({item["city"] for item in records if item["city"]}),
        "outputs": [str(csv_path), str(json_path), str(xlsx_path)],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
